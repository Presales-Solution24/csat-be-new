from flask import Blueprint, request, jsonify, send_file
from app.models import db, Project, Task, ProductType
from datetime import datetime
import secrets
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from io import BytesIO
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import quote_sheetname

project_bp = Blueprint('project_bp', __name__)

@project_bp.route('/project-satisfaction', methods=['POST'])
def save_project_satisfaction():
    data = request.get_json()
    project_id = data.get('project_id')

    try:
        if project_id:
            project = Project.query.get(project_id)
            if not project:
                return jsonify({'error': 'Project not found'}), 404
            # Update existing project
            project.nama_company = data.get('namaCompany')
            project.nama_pribadi = data.get('namaPribadi')
            project.email = data.get('email')
            project.no_hp = data.get('noHp')
            project.jabatan = data.get('jabatan')
            project.project_title = data.get('projectTitle')
            project.nama_presales = data.get('namaPresales')
            project.tanggal_projek = data.get('tanggalProjek')
            project.tanggal_approved = data.get('tanggalApproved')
            project.approved = data.get('approved')
            project.rating = data.get('rating')
            project.tipe_produk = data.get('tipeProduk')
            project.kategori_produk = data.get('kategoriProduk')

            # Jika token belum ada, generate
            if not project.scoring_token:
                project.scoring_token = secrets.token_urlsafe(32)

            # Hapus semua task lama
            Task.query.filter_by(project_id=project_id).delete()
        else:
            # Buat project baru
            project = Project(
                nama_company=data.get('namaCompany'),
                nama_pribadi=data.get('namaPribadi'),
                email=data.get('email'),
                no_hp=data.get('noHp'),
                jabatan=data.get('jabatan'),
                project_title=data.get('projectTitle'),
                nama_presales=data.get('namaPresales'),
                tanggal_projek=data.get('tanggalProjek'),
                tanggal_approved=data.get('tanggalApproved'),
                approved=data.get('approved'),
                rating=data.get('rating'),
                tipe_produk=data.get('tipeProduk'),
                kategori_produk=data.get('kategoriProduk'),
                scoring_token=secrets.token_urlsafe(32)  # generate token
            )
            db.session.add(project)
            db.session.flush()  # agar project.id langsung tersedia

        # Simpan task-task baru
        tasks = data.get('tasks', [])
        for t in tasks:
            task = Task(
                project_id=project.id,
                task=t.get('task'),
                expected_result=t.get('expectedResult'),
                actual_result=t.get('actualResult')
            )
            db.session.add(task)

        db.session.commit()
        return jsonify({"message": "Project saved", "project_id": project.id}), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@project_bp.route('/project-satisfaction/<int:id>', methods=['GET'])
def get_project_satisfaction(id):
    project = Project.query.get(id)
    if not project:
        return jsonify({"error": "Project not found"}), 404

    tasks = Task.query.filter_by(project_id=id).all()

    return jsonify({
        "namaCompany": project.nama_company,
        "namaPribadi": project.nama_pribadi,
        "email": project.email,
        "noHp": project.no_hp,
        "jabatan": project.jabatan,
        "projectTitle": project.project_title,
        "namaPresales": project.nama_presales,
        "tanggalProjek": str(project.tanggal_projek),
        "tanggalApproved": str(project.tanggal_approved),
        "approved": project.approved,
        "rating": project.rating,
        "tipeProduk": project.tipe_produk,
        "kategoriProduk": project.kategori_produk,
        "scoring_token": project.scoring_token,
        "tasks": [
            {
                "task": t.task,
                "expectedResult": t.expected_result,
                "actualResult": t.actual_result
            }
            for t in tasks
        ]
    })


@project_bp.route('/project-satisfaction', methods=['GET'])
def list_project_satisfaction():
    projects = Project.query.order_by(Project.id.desc()).all()
    result = []
    for project in projects:
        tasks = Task.query.filter_by(project_id=project.id).all()
        result.append({
            "id": project.id,
            "nama_company": project.nama_company,
            "nama_pribadi": project.nama_pribadi,
            "email": project.email,
            "no_hp": project.no_hp,
            "jabatan": project.jabatan,
            "project_title": project.project_title,
            "nama_presales": project.nama_presales,
            "tanggal_projek": str(project.tanggal_projek) if project.tanggal_projek else None,
            "tanggal_approved": str(project.tanggal_approved) if project.tanggal_approved else None,
            "approved": project.approved,
            "rating": project.rating,
            "tipe_produk": project.tipe_produk,
            "kategori_produk": project.kategori_produk,
            "tasks": [
                {
                    "id": task.id,
                    "project_id": task.project_id,
                    "task_name": task.task,
                    "expected_result": task.expected_result,
                    "actual_result": task.actual_result
                } for task in tasks
            ]
        })
    return jsonify(result)


@project_bp.route('/project-satisfaction/<int:id>', methods=['DELETE'])
def delete_project_satisfaction(id):
    try:
        project = Project.query.get(id)
        if not project:
            return jsonify({"error": "Project not found"}), 404

        # Hapus semua task terkait
        Task.query.filter_by(project_id=id).delete()

        # Hapus project
        db.session.delete(project)
        db.session.commit()

        return jsonify({"message": "Project and related tasks deleted successfully."}), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@project_bp.route('/public-scoring/<token>', methods=['GET'])
def get_project_by_token(token):
    project = Project.query.filter_by(scoring_token=token).first()
    if not project:
        return jsonify({"error": "Invalid or expired token"}), 404

    tasks = Task.query.filter_by(project_id=project.id).all()

    return jsonify({
        "id": project.id,
        "nama_company": project.nama_company,
        "nama_pribadi": project.nama_pribadi,
        "email": project.email,
        "no_hp": project.no_hp,
        "jabatan": project.jabatan,
        "project_title": project.project_title,
        "nama_presales": project.nama_presales,
        "tanggal_projek": project.tanggal_projek,
        "tanggal_approved": project.tanggal_approved,
        "approved": project.approved,
        "rating": project.rating,
        "tipe_produk": project.tipe_produk,
        "kategori_produk": project.kategori_produk,
        "tasks": [
            {
                "task": t.task,
                "expected_result": t.expected_result,
                "actual_result": t.actual_result
            }
            for t in tasks
        ]
    })

@project_bp.route('/public-scoring/<token>', methods=['POST'])
def update_project_scoring_by_token(token):
    project = Project.query.filter_by(scoring_token=token).first()
    if not project:
        return jsonify({"error": "Invalid or expired token"}), 404

    data = request.get_json()
    project.tanggal_approved = data.get("tanggal_approved", "")
    project.approved = data.get("approved", "")
    project.rating = data.get("rating", "")

    db.session.commit()
    return jsonify({"message": "Scoring updated successfully"})

@project_bp.route('/project-satisfaction/export/excel', methods=['GET'])
def export_project_satisfaction_excel():
    projects = Project.query.order_by(Project.id.desc()).all()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Project Satisfaction"

    # Header
    headers = [
        "ID", "Nama Company", "Nama Pribadi", "Email", "No HP", "Jabatan",
        "Project Title", "Nama Presales", "Tanggal Projek", "Tanggal Approved",
        "Approved", "Rating", "Tipe Produk", "Kategori Produk",
        "Task 1", "Expected 1", "Actual 1",
        "Task 2", "Expected 2", "Actual 2",
        "Task 3", "Expected 3", "Actual 3"
    ]
    sheet.append(headers)

    for project in projects:
        tasks = Task.query.filter_by(project_id=project.id).all()
        row = [
            project.id,
            project.nama_company,
            project.nama_pribadi,
            project.email,
            project.no_hp,
            project.jabatan,
            project.project_title,
            project.nama_presales,
            str(project.tanggal_projek) if project.tanggal_projek else "",
            str(project.tanggal_approved) if project.tanggal_approved else "",
            "Yes" if project.approved else "No",
            project.rating,
            project.tipe_produk,
            project.kategori_produk,
        ]

        # Tambahkan hingga 3 task
        for i in range(3):
            if i < len(tasks):
                row.extend([
                    tasks[i].task,
                    tasks[i].expected_result,
                    tasks[i].actual_result,
                ])
            else:
                row.extend(["", "", ""])

        sheet.append(row)

    # Autofit column width
    for col in sheet.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        sheet.column_dimensions[column].width = max_length + 2

    # Simpan ke dalam memory
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)

    return send_file(
        stream,
        as_attachment=True,
        download_name="Project_Satisfaction_List.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@project_bp.route("/project-satisfaction/template/excel", methods=["GET"])
def download_project_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "ProjectSatisfaction"

    headers = [
        "Tanggal Project", "Nama Company", "Nama Pribadi", "Email", "Judul Project",
        "Nama Presales", "Tipe Produk", "Kategori Produk",
        "Task 1", "Expected 1", "Actual 1",
        "Task 2", "Expected 2", "Actual 2",
        "Task 3", "Expected 3", "Actual 3",
        "Tanggal Approve", "Approved", "Rating"
    ]
    ws.append(headers)

    # Data contoh
    product_types = ProductType.query.all()
    first_type = product_types[0] if product_types else None

    example_row = [
        "2025-08-01", "PT Contoh Abadi", "Budi Santoso", "budi@example.com", "Digitalisasi Arsip",
        "Guntur", 
        first_type.tipe_produk if first_type else "", 
        first_type.kategori_produk if first_type else "",
        "Instalasi Scanner", "Berhasil scanning 1000 dokumen/hari", "Tercapai",
        "Pelatihan", "User memahami fungsi-fungsi utama", "Tercapai",
        "Integrasi sistem", "Terhubung ke DMS", "Tercapai",
        "2025-08-05", "Yes", "5"
    ]
    ws.append(example_row)

    # Sheet untuk dropdown
    dropdown_sheet = wb.create_sheet("Dropdown")
    dropdown_sheet.append(["Tipe Produk", "Kategori Produk"])
    for pt in product_types:
        dropdown_sheet.append([pt.tipe_produk, pt.kategori_produk])

    last_row = len(product_types) + 1
    tipe_range = f"Dropdown!$A$2:$A${last_row}"

    # Named range
    named_range = DefinedName(name="TipeProdukList", attr_text=tipe_range)
    wb.defined_names.add(named_range)

    # Data validation: Tipe Produk
    dv_tipe = DataValidation(
        type="list",
        formula1="=TipeProdukList",
        allow_blank=True,
        showDropDown=True
    )
    ws.add_data_validation(dv_tipe)
    dv_tipe.add("G3:G100")  # baris ke-3 ke bawah (bukan baris contoh)

    # VLOOKUP untuk Kategori Produk
    for row in range(3, 101):
        ws[f"H{row}"] = f'=IFERROR(VLOOKUP(G{row},Dropdown!A:B,2,FALSE),"")'

    # Data validation: Approved (Yes/No)
    dv_approved = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=True
    )
    ws.add_data_validation(dv_approved)
    dv_approved.add("R3:R100")

    # Simpan ke memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="project_satisfaction_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
@project_bp.route("/project-satisfaction/upload/preview", methods=["POST"])
def preview_excel():
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files['file']
    if file.filename == "":
        return jsonify({"error": "No selected file"}), 400

    try:
        df = pd.read_excel(file)
        df.columns = df.columns.str.strip()

        required_columns = [
            "Tanggal Project", "Nama Company", "Nama Pribadi", "Email",
            "Judul Project", "Nama Presales", "Tipe Produk", "Kategori Produk",
            "Task 1", "Expected 1", "Actual 1",
            "Task 2", "Expected 2", "Actual 2",
            "Task 3", "Expected 3", "Actual 3",
            "Tanggal Approve", "Approved", "Rating"
        ]

        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return jsonify({"error": f"Missing columns: {', '.join(missing_columns)}"}), 400

        # Ambil semua pasangan tipe-kategori dari DB
        type_map = {
            pt.tipe_produk.strip(): pt.kategori_produk.strip()
            for pt in db.session.query(ProductType).all()
        }

        # Validasi dan tambahkan status per baris
        preview_data = []
        for _, row in df.iterrows():
            row_dict = row.fillna("").to_dict()
            tipe = row_dict.get("Tipe Produk", "").strip()
            kategori = row_dict.get("Kategori Produk", "").strip()
            expected_kategori = type_map.get(tipe)

            if expected_kategori and kategori == expected_kategori:
                row_dict["status"] = "valid"
            else:
                row_dict["status"] = "invalid"

            preview_data.append(row_dict)

        return jsonify({"preview": preview_data}), 200

    except Exception as e:
        return jsonify({"error": f"Failed to process file: {str(e)}"}), 500

@project_bp.route("/project-satisfaction/upload/save", methods=["POST"])
def save_excel():
    data = request.json
    if not data:
        return jsonify({"error": "No data provided"}), 400

    try:
        for item in data:
            project = Project(
                tanggal_projek=item.get("Tanggal Project"),
                nama_company=item.get("Nama Company"),
                nama_pribadi=item.get("Nama Pribadi"),
                email=item.get("Email"),
                project_title=item.get("Judul Project"),
                nama_presales=item.get("Nama Presales"),
                tipe_produk=item.get("Tipe Produk"),
                kategori_produk=item.get("Kategori Produk"),
                tanggal_approved=item.get("Tanggal Approve"),
                approved = item.get("Approved") == "Yes",
                rating=item.get("Rating"),
                scoring_token=secrets.token_urlsafe(32),
            )
            db.session.add(project)
            db.session.flush()

            # Tambah task jika ada
            for i in range(1, 4):
                task = item.get(f"Task {i}")
                expected = item.get(f"Expected {i}")
                actual = item.get(f"Actual {i}")
                if task or expected or actual:
                    db.session.add(Task(
                        project_id=project.id,
                        task=task,
                        expected_result=expected,
                        actual_result=actual
                    ))

        db.session.commit()
        return jsonify({"message": "Data berhasil disimpan"}), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Gagal menyimpan data: {str(e)}"}), 500
